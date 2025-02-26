import os
from openpyxl import load_workbook
from django.core.management.base import BaseCommand
from django.utils.text import slugify
from django.contrib.auth import get_user_model
from topeducation.models import Blog, CategoriaBlog, Autor

User = get_user_model()

class Command(BaseCommand):
    help = 'Importa blogs desde Excel con contenido HTML formateado'

    def handle(self, *args, **options):
        excel_path = "C:/Users/felip/Desktop/Top Education Archivo Maestro 2025.xlsx"
        
        try:
            wb = load_workbook(excel_path)
            ws = wb.active
            column_mapping = {cell.value.strip(): idx for idx, cell in enumerate(ws[1], start=0) if cell.value}
            print(column_mapping)
            
            required_columns = ["Titulo", "Meta descripción", "Palabra clave", "Autor", "Categoría", "Objetivo", "Miniatura"]
            for col in required_columns:
                if col not in column_mapping:
                    self.stdout.write(self.style.ERROR(f'❌ Columna faltante: {col}'))
                    return

        except Exception as e:
            self.stdout.write(self.style.ERROR(f'❌ Error cargando Excel: {str(e)}'))
            return

        # Contenido HTML formateado (el que está arriba)
        contenido_html = """<h1>Guía: cómo encontrar trabajo con poca experiencia</h1>
 
 <p>En un mundo laboral cada vez más competitivo, encontrar trabajo con poca experiencia puede resultar desafiante.</p>
 
 <p>Sin embargo, con la combinación adecuada de estrategias, enfoque y determinación, es posible abrirse camino hacia oportunidades laborales gratificantes y relevantes; esta guía está diseñada para proporcionar consejos prácticos y recursos útiles para aquellos que buscan ingresar al mercado laboral.</p>
 
 <img src="/assets/Piezas/como_encontrar_trabajo_con_poca_experiencia_miniatura.png" alt="Como encontrar trabajo con poca experiencia"/>
    
<h2>Importancia de la experiencia laboral</h2>
 
 <p>La experiencia laboral es un activo invaluable que los empleadores valoran profundamente al evaluar candidatos.</p>
 
 <p>No obstante, la noción de "trabajo con poca experiencia" no debería desalentar a los buscadores de empleo, ya que la experiencia va más allá de la cantidad de años en el campo laboral.</p>
 
 <p>Los empleadores buscan evidencia de habilidades prácticas, capacidades de resolución de problemas y adaptabilidad, los empleadores también reconocen el valor del potencial y la motivación intrínseca en los candidatos con poca experiencia.</p>
 
 <div class="cta">
    <a href="https://info.top.education/es-mx/como-encontrar-trabajo-con-poca-experiencia" target="_blank"></a>
 </div>
 
 <h2>Consejos para elaborar un currículum impactante</h2>
 
 <p>Cuando se trata de crear un currículum efectivo para aquellos que buscan trabajo con poca experiencia, es fundamental destacar tanto las habilidades como los logros relevantes de manera estratégica.</p>
 
 <p>La estructura de tu currículum debe incluir secciones bien definidas, como un resumen profesional, experiencia laboral (así sea limitada), educación, habilidades y logros.</p>
 
 <p>En lugar de centrarte únicamente en la experiencia laboral previa, destaca tus habilidades transferibles y logros significativos; utiliza frases concisas y poderosas para describir cómo estos han contribuido de manera significativa, incluso con una experiencia limitada.</p>
 
 <h2>¿Cómo prepararse para la entrevista?</h2>
 
 <p>Cuando te enfrentes a preguntas sobre tu experiencia laboral durante una entrevista, es esencial destacar tus habilidades, en lugar de centrarte en tu poca experiencia laboral, resalta tus experiencias académicas o proyectos que demuestren capacidades relevantes para el puesto al que aspiras.</p>
 
 <p>Utiliza ejemplos concretos para ilustrar cómo has aplicado estas habilidades en situaciones prácticas y cómo estás preparado para enfrentar los desafíos del nuevo rol.</p>
 
 <p>Al demostrar tu motivación, capacidad de aprendizaje y entusiasmo por crecer profesionalmente, podrás compensar la falta de experiencia laboral directa y presentarte como un candidato valioso y comprometido.</p>
 
 <h2>Redes profesionales y su importancia</h2>
 
 <p>Construir y aprovechar redes profesionales es fundamental para avanzar en tu carrera, incluso si tienes poca experiencia.</p>
 
 <ul>
     <li>Comienza por establecer perfiles en plataformas profesionales como LinkedIn, donde puedes conectar con personas de tu industria y participar en grupos de discusión relevantes.</li>
     <li>No subestimes el poder de las relaciones personales; asiste a eventos de networking, conferencias o charlas profesionales donde puedas conocer a otros en tu campo.</li>
     <li>Sé proactivo al buscar mentores o profesionales con más experiencia que estén dispuestos a guiarte y brindarte consejos.</li>
     <li>Además, no tengas miedo de pedir información o consejos a personas que admires en tu industria.</li>
 </ul>
 
 <p>Recuerda que cada conexión que hagas puede abrir nuevas oportunidades y ofrecer valiosos conocimientos, incluso en las etapas iniciales de tu carrera.</p>
 
 <h2>Estrategias para conseguir experiencia práctica</h2>
 
 <p>Cuando te encuentras en la situación de querer ganar experiencia práctica, pero careces de experiencia laboral previa, existen diversas estrategias que puedes emplear para construir un historial sólido.</p>
 
 <p>Una opción es buscar oportunidades de voluntariado en organizaciones sin fines de lucro con tu campo de interés; otra alternativa es realizar proyectos independientes o freelancing, ofreciendo tus servicios en plataformas en línea o a través de tu red de contactos.</p>
 
 <p>Al combinar estas estrategias con una actitud proactiva y un enfoque en el aprendizaje continuo, podrás desarrollar una base sólida de experiencia práctica, incluso cuando partas desde cero en términos de experiencia laboral.</p>
 
 <p>Si bien la experiencia tradicional puede ser un factor determinante, lo que abre oportunidades es estar dispuestos a demostrar valía y compromiso en el proceso de selección.</p>
 
 <img src="/assets/Piezas/como_encontrar_trabajo_con_poca_experiencia_vertical.png" alt="Como encontrar trabajo con poca experiencia">
 
 <h2>Conclusión</h2>
 
 <p>La búsqueda de empleo puede ser desafiante, especialmente cuando se tiene poca experiencia, sin embargo, con las estrategias adecuadas y una mentalidad perseverante, es posible alcanzar el éxito.</p>
 
 <p>A lo largo de esta guía, hemos proporcionado consejos prácticos, recursos útiles y orientación valiosa para ayudarte a navegar por este proceso con confianza y determinación.</p>
 
 <p>Pero este es solo el comienzo de tu viaje hacia una carrera gratificante, para obtener aún más información detallada, herramientas y ejemplos prácticos que te ayudarán a destacarte en tu búsqueda de empleo, te invitamos a descargar nuestra guía completa.</p>
 
 <p>¡No dejes pasar esta oportunidad para impulsar tu carrera hacia el éxito!</p>
 
 <div class="cta">
    <a href="https://info.top.education/es-mx/como-encontrar-trabajo-con-poca-experiencia" target="_blank">CTA Cómo encontrar trabajo con poca experiencia</a>
</div>


"""

        row = next(ws.iter_rows(min_row=2, max_row=2))
        try:
            blog_data = self.obtener_datos_fila(row, column_mapping)
            if blog_data['titulo']:
                self.guardar_blog(blog_data, contenido_html, 3)
        except Exception as e:
            self.stdout.write(self.style.ERROR(f'❌ Error procesando blog: {str(e)}'))

        self.stdout.write(self.style.SUCCESS('\n✅ Proceso completado'))

    def obtener_datos_fila(self, row, column_mapping):
        return {
            'titulo': row[column_mapping["Titulo"]].value or '',
            'meta_desc': row[column_mapping["Meta descripción"]].value or '',
            'palabra_clave': row[column_mapping["Palabra clave"]].value or '',
            'autor': row[column_mapping["Autor"]].value or '',
            'categoria': row[column_mapping["Categoría"]].value or 'General',
            'objetivo': row[column_mapping["Objetivo"]].value or '',
            'miniatura': row[column_mapping["Miniatura"]].value or ''
        }

    def guardar_blog(self, data, contenido, idx):
        try:
            
            counter = 1
            
    

            categoria, _ = CategoriaBlog.objects.get_or_create(nombre_categoria_blog=data['categoria'])
            autor = Autor.objects.filter(nombre_autor=data['autor']).first() or User.objects.first()
            #print(data['Miniatura'])

            Blog.objects.create(
                nombre_blog=data['titulo'],
                metadescripcion_blog=data['meta_desc'],
                palabra_clave_blog=data['palabra_clave'],
                autor_blog=autor,
                categoria_blog=categoria,
                objetivo_blog=data['objetivo'],
                contenido=contenido,
                miniatura_blog = data['miniatura']
            )

            self.stdout.write(self.style.SUCCESS(f"✅ Blog guardado - {data['titulo']}"))

        except Exception as e:
            self.stdout.write(self.style.ERROR(f'❌ Error guardando blog: {str(e)}'))